"""
Excel parser for Sid Sloth's AI Tech Stocks Portfolio spreadsheet.

Only processes the first sheet ("AI Tech Stocks Port - Current").
All other sheets (dated snapshots, Investing Guide, etc.) are ignored.

Known column layout:
    B  → stock name          H  → ticker (may be empty — fallback to B)
    I  → buy_high (From)     J  → buy_low (To)
    L  → sell_start (1st trim)
    Z  → pe_current          AD → pe_range_high     AE → pe_range_low
    AI → fair_value           V → egf
    W  → egf_direction        X → egf_12m
    Y  → fundamentals        AH → trend_status
    T  → prob_new_ath        AF → strategy_text
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

DATA_START_ROW = 6

# Section labels that mark category boundaries (column A, case-insensitive match)
_SECTION_LABELS = {
    "PRIMARY", "SECONDARY", "SLEEPERS", "TECH FUNDS",
    "WAR WITH CHINA MEGA-TREND", "GLOBAL WARMING MEGA-TREND",
    "MEDIUM RISK STOCKS", "HEALTHCARE", "HOUSING STOCKS",
    "HIGH RISK STOCKS", "CRYPTOS",
}

# Rows to skip entirely
_SKIP_PATTERNS = (
    "TARGET", "Av Exposure", "S&P 500", "Disclaimer",
    "Stocks Mentioned", "Ad hoc analysis",
)


@dataclass
class StockRow:
    """One ticker's data extracted from the main sheet."""
    stock_name: str
    ticker: str
    buy_high: float | None = None
    buy_low: float | None = None
    sell_start: float | None = None
    pe_range_high: float | None = None
    pe_range_low: float | None = None
    fair_value: float | None = None
    egf: float | None = None
    egf_direction: float | None = None
    egf_12m: float | None = None
    fundamentals: float | None = None
    trend_status: str | None = None
    prob_new_ath: float | None = None
    strategy_text: str | None = None
    category: str = "PRIMARY"
    stock_type: str | None = None  # Range Trader / Trender / etc.


@dataclass
class SheetSnapshot:
    """All tickers from the current sheet."""
    sheet_name: str
    stocks: list[StockRow] = field(default_factory=list)


@dataclass
class PrincipleEntry:
    """Placeholder — principles no longer parsed from Excel."""
    section_number: str
    title: str
    content: str
    category: str


def _safe_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# Known display name → ticker overrides for stocks where the Excel doesn't have
# the correct ticker symbol in any column
_TICKER_OVERRIDES = {
    "circle": "CRCL",
    "bitmine": "BTCM",
    "bitmine - eth": "BTCM",
}

# Name overrides: when the Excel has uninformative names for certain tickers
_NAME_OVERRIDES = {
    "MSTR": "MicroStrategy",
    "BTCM": "Bitmine",
}


def _clean_ticker(raw: str) -> str:
    """Extract a clean ticker symbol from messy Excel values.

    Examples:
        'GOOG 0HD6 YS' → 'GOOG'
        'META 0QZI - YS' → 'META'
        'MU 0R2T Y' → 'MU'
        'IBM.L YS' → 'IBM'
        'BIDU 9888' → 'BIDU'
        'COIN I' → 'COIN'
        'SMT - L SI' → 'SMT'
        'WTAI.L / INTL.L IS' → 'WTAI'
        'LMT - 0R3E.L' → 'LMT'
    """
    # Check overrides first
    override = _TICKER_OVERRIDES.get(raw.lower().strip())
    if override:
        return override

    # Take first space-separated token
    token = raw.split()[0] if raw else raw
    # Strip exchange suffixes (.L, .NV, etc.) — keep the base symbol
    if "." in token:
        base = token.split(".")[0]
        suffix = token.split(".", 1)[1]
        if len(suffix) <= 3:
            token = base
    return token.upper()


def _is_section_label(text: str) -> bool:
    """Check if text is a section category label."""
    return text.upper().strip() in _SECTION_LABELS


def _should_skip(text: str) -> bool:
    """Check if row should be skipped (summary rows, disclaimers, etc.)."""
    return any(text.startswith(p) for p in _SKIP_PATTERNS)


def parse_stock_sheet(ws: Worksheet, sheet_name: str) -> SheetSnapshot:
    """Parse the current stock sheet into a SheetSnapshot."""
    snapshot = SheetSnapshot(sheet_name=sheet_name)
    current_category = "PRIMARY"

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        cells = {}
        for cell in row:
            if cell.value is not None:
                cells[cell.column_letter] = cell.value

        a_val = _safe_str(cells.get("A"))
        b_val = _safe_str(cells.get("B"))
        h_val = _safe_str(cells.get("H"))

        # Check for section label in column A
        if a_val and _is_section_label(a_val):
            current_category = a_val.upper().strip()
            continue

        # Skip empty rows
        if not a_val and not b_val and not h_val:
            continue

        # Skip summary/disclaimer rows
        if b_val and _should_skip(b_val):
            continue
        if a_val and _should_skip(a_val):
            continue

        # Determine stock name and ticker
        # Normal case: B=stock name, H=ticker
        # Fallback: when H is empty, B might contain the ticker (e.g. rows 45-51)
        stock_name = b_val
        raw_ticker = h_val

        if not raw_ticker and b_val:
            # H is empty — B might contain ticker or name
            # First check if b_val has a known override (e.g. "Circle" → CRCL)
            if b_val.lower().strip() in _TICKER_OVERRIDES:
                raw_ticker = b_val
                stock_name = b_val  # Use original name
            else:
                candidate = b_val.split()[0].strip()
                if len(candidate) <= 6 and candidate.replace(".", "").isalpha():
                    raw_ticker = candidate
                    stock_name = b_val  # Keep full B value as name
                    # If A has a descriptive label, use it as name instead
                    if a_val:
                        try:
                            float(a_val)
                        except (ValueError, TypeError):
                            if not _is_section_label(a_val):
                                stock_name = a_val

        if not raw_ticker:
            continue
        if not stock_name:
            stock_name = raw_ticker

        # Skip if ticker has no letters (crypto numeric rows, dates, etc.)
        if not any(c.isalpha() for c in raw_ticker):
            continue

        # Skip rows where H contains header text like "Ticker" or "How to X10?"
        if raw_ticker.lower().startswith("ticker") or raw_ticker.lower().startswith("how to"):
            continue

        # Clean the ticker
        ticker = _clean_ticker(raw_ticker)

        # Apply name overrides
        if ticker in _NAME_OVERRIDES:
            stock_name = _NAME_OVERRIDES[ticker]

        # Skip "deceased" / "R.I.P." stocks
        if a_val and a_val.lower() in ("deceased", "r.i.p.", "to be removed"):
            logger.info("Skipping %s (%s) — marked as %s", ticker, stock_name, a_val)
            continue

        # Determine stock_type from column A if it's not a section label or number
        stock_type = None
        if a_val:
            try:
                float(a_val)
            except (ValueError, TypeError):
                if not _is_section_label(a_val) and a_val.lower() not in ("deceased", "r.i.p.", "to be removed"):
                    stock_type = a_val  # e.g. "Range Trader", "Trender", "Wind Power"

        stock = StockRow(
            stock_name=stock_name,
            ticker=ticker,
            buy_high=_safe_float(cells.get("I")),
            buy_low=_safe_float(cells.get("J")),
            sell_start=_safe_float(cells.get("L")),
            pe_range_high=_safe_float(cells.get("AD")),
            pe_range_low=_safe_float(cells.get("AE")),
            fair_value=_safe_float(cells.get("AI")),
            egf=_safe_float(cells.get("V")),
            egf_direction=_safe_float(cells.get("W")),
            egf_12m=_safe_float(cells.get("X")),
            fundamentals=_safe_float(cells.get("Y")),
            trend_status=_safe_str(cells.get("AH")),
            prob_new_ath=_safe_float(cells.get("T")),
            strategy_text=_safe_str(cells.get("AF")),
            category=current_category,
            stock_type=stock_type,
        )
        # Skip stocks with no useful data at all
        has_data = any([
            stock.buy_high, stock.buy_low, stock.sell_start,
            stock.fair_value, stock.strategy_text, stock.trend_status,
        ])
        if not has_data:
            logger.debug("Skipping %s (%s) — no price/strategy data", ticker, stock_name)
            continue

        snapshot.stocks.append(stock)

    logger.info("Parsed %d stocks from sheet '%s'", len(snapshot.stocks), sheet_name)
    return snapshot


def parse_workbook(file_path: str | Path) -> dict:
    """
    Parse the first sheet of the workbook only.

    Returns:
        {
            "current": SheetSnapshot,
            "dated_snapshots": [],
            "principles": [],
            "diffs": [],
        }
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)

    result = {
        "current": None,
        "dated_snapshots": [],
        "principles": [],
        "diffs": [],
    }

    # Parse only the first sheet (current portfolio)
    current_sheet_name = "AI Tech Stocks Port - Current"
    if current_sheet_name in wb.sheetnames:
        result["current"] = parse_stock_sheet(wb[current_sheet_name], current_sheet_name)
    elif wb.sheetnames:
        # Fallback to first sheet
        first = wb.sheetnames[0]
        logger.warning("Expected sheet '%s' not found, using '%s'", current_sheet_name, first)
        result["current"] = parse_stock_sheet(wb[first], first)

    wb.close()
    stock_count = len(result["current"].stocks) if result["current"] else 0
    logger.info("Workbook parsed: %d stocks from first sheet only", stock_count)
    return result
